from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime

from .models import Vehiculo, Conductor, UnidadSolicitante, ActividadSalud, Destino, ReservaMovil
from .forms import VehiculoForm, ConductorForm, UnidadSolicitanteForm, ActividadSaludForm, DestinoForm, ReservaMovilForm


# -------------------------
# FUNCIONES HELPER PARA FILTRADO
# -------------------------

def aplicar_filtro_vehiculos(queryset, busqueda):
    if busqueda:
        return queryset.filter(
            patente__icontains=busqueda
        ) | queryset.filter(
            tipo__icontains=busqueda
        ) | queryset.filter(
            marca__icontains=busqueda
        ) | queryset.filter(
            modelo__icontains=busqueda
        ) | queryset.filter(
            estado__icontains=busqueda
        )
    return queryset


def aplicar_filtro_conductores(queryset, busqueda):
    if busqueda:
        return queryset.filter(
            nombre_completo__icontains=busqueda
        ) | queryset.filter(
            rut__icontains=busqueda
        ) | queryset.filter(
            tipo_licencia__icontains=busqueda
        ) | queryset.filter(
            estado__icontains=busqueda
        )
    return queryset


def aplicar_filtro_unidades(queryset, busqueda):
    if busqueda:
        return queryset.filter(
            nombre__icontains=busqueda
        ) | queryset.filter(
            responsable__icontains=busqueda
        ) | queryset.filter(
            correo_contacto__icontains=busqueda
        )
    return queryset


def aplicar_filtro_actividades(queryset, busqueda):
    if busqueda:
        return queryset.filter(
            nombre__icontains=busqueda
        ) | queryset.filter(
            descripcion__icontains=busqueda
        )
    return queryset


def aplicar_filtro_destinos(queryset, busqueda):
    if busqueda:
        return queryset.filter(
            nombre_lugar__icontains=busqueda
        ) | queryset.filter(
            direccion__icontains=busqueda
        ) | queryset.filter(
            referencia__icontains=busqueda
        )
    return queryset


def aplicar_filtro_reservas(queryset, busqueda):
    if busqueda:
        return queryset.filter(
            unidad_solicitante__nombre__icontains=busqueda
        ) | queryset.filter(
            actividad__nombre__icontains=busqueda
        ) | queryset.filter(
            destino__nombre_lugar__icontains=busqueda
        ) | queryset.filter(
            estado__icontains=busqueda
        )
    return queryset


def notificar_errores_formulario(request, form):
    errores = []

    for campo, lista_errores in form.errors.items():
        etiqueta = form.fields[campo].label if campo in form.fields else 'Formulario'

        for error in lista_errores:
            errores.append(f'{etiqueta}: {error}')

    if errores:
        messages.error(request, 'Revise el formulario. ' + ' '.join(errores))


def inicio(request):
    total_vehiculos = Vehiculo.objects.count()
    total_conductores = Conductor.objects.count()
    total_unidades = UnidadSolicitante.objects.count()
    total_actividades = ActividadSalud.objects.count()
    total_destinos = Destino.objects.count()
    total_reservas = ReservaMovil.objects.count()

    context = {
        'total_vehiculos': total_vehiculos,
        'total_conductores': total_conductores,
        'total_unidades': total_unidades,
        'total_actividades': total_actividades,
        'total_destinos': total_destinos,
        'total_reservas': total_reservas,
    }

    return render(request, 'inicio.html', context)


# -------------------------
# CRUD VEHICULO
# -------------------------

def listar_vehiculos(request):
    busqueda = request.GET.get('buscar', '')

    vehiculos = Vehiculo.objects.all().order_by('id')
    vehiculos = aplicar_filtro_vehiculos(vehiculos, busqueda)

    paginator = Paginator(vehiculos, 1)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'vehiculos/listar.html', {
        'page_obj': page_obj,
        'busqueda': busqueda
    })

def detalle_vehiculo(request, id):
    vehiculo = get_object_or_404(Vehiculo, id=id)
    return render(request, 'vehiculos/detalle.html', {
        'vehiculo': vehiculo
    })


def crear_vehiculo(request):
    form = VehiculoForm(request.POST or None, request.FILES or None)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, 'Vehículo registrado correctamente.')
            return redirect('listar_vehiculos')
        notificar_errores_formulario(request, form)

    return render(request, 'vehiculos/formulario.html', {
        'form': form,
        'titulo': 'Registrar vehículo'
    })


def editar_vehiculo(request, id):
    vehiculo = get_object_or_404(Vehiculo, id=id)
    form = VehiculoForm(request.POST or None, request.FILES or None, instance=vehiculo)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, 'Vehículo actualizado correctamente.')
            return redirect('listar_vehiculos')
        notificar_errores_formulario(request, form)

    return render(request, 'vehiculos/formulario.html', {
        'form': form,
        'titulo': 'Editar vehículo'
    })


def eliminar_vehiculo(request, id):
    vehiculo = get_object_or_404(Vehiculo, id=id)

    if request.method == 'POST':
        # Validar si el vehículo tiene reservas asociadas
        reservas_asociadas = ReservaMovil.objects.filter(vehiculo=vehiculo)
        
        if reservas_asociadas.exists():
            messages.error(
                request, 
                f'No se puede eliminar el vehículo "{vehiculo.patente}" '
                f'porque tiene {reservas_asociadas.count()} reserva(s) de móvil asociada(s). '
                'Debe cancelar o finalizar todas las reservas antes de eliminar.'
            )
            return redirect('listar_vehiculos')
        
        vehiculo.delete()
        messages.success(request, 'Vehículo eliminado correctamente.')
        return redirect('listar_vehiculos')

    return redirect('listar_vehiculos')


# -------------------------
# CRUD CONDUCTOR
# -------------------------

def listar_conductores(request):
    busqueda = request.GET.get('buscar', '')

    conductores = Conductor.objects.all().order_by('id')
    conductores = aplicar_filtro_conductores(conductores, busqueda)

    paginator = Paginator(conductores, 1)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'conductores/listar.html', {
        'page_obj': page_obj,
        'busqueda': busqueda
    })

def detalle_conductor(request, id):
    conductor = get_object_or_404(Conductor, id=id)

    return render(request, 'conductores/detalle.html', {
        'conductor': conductor
    })


def crear_conductor(request):
    form = ConductorForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, 'Conductor registrado correctamente.')
            return redirect('listar_conductores')
        notificar_errores_formulario(request, form)

    return render(request, 'conductores/formulario.html', {
        'form': form,
        'titulo': 'Registrar conductor'
    })


def editar_conductor(request, id):
    conductor = get_object_or_404(Conductor, id=id)
    form = ConductorForm(request.POST or None, instance=conductor)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, 'Conductor actualizado correctamente.')
            return redirect('listar_conductores')
        notificar_errores_formulario(request, form)

    return render(request, 'conductores/formulario.html', {
        'form': form,
        'titulo': 'Editar conductor'
    })


def eliminar_conductor(request, id):
    conductor = get_object_or_404(Conductor, id=id)

    if request.method == 'POST':
        # Validar si el conductor tiene reservas asociadas
        reservas_asociadas = ReservaMovil.objects.filter(conductor=conductor)
        
        if reservas_asociadas.exists():
            messages.error(
                request, 
                f'No se puede eliminar el conductor "{conductor.nombre_completo}" '
                f'porque tiene {reservas_asociadas.count()} reserva(s) de móvil asociada(s). '
                'Debe cancelar o finalizar todas las reservas antes de eliminar.'
            )
            return redirect('listar_conductores')
        
        conductor.delete()
        messages.success(request, 'Conductor eliminado correctamente.')
        return redirect('listar_conductores')

    return redirect('listar_conductores')


# -------------------------
# CRUD UNIDAD SOLICITANTE
# -------------------------

def listar_unidades(request):
    busqueda = request.GET.get('buscar', '')

    unidades = UnidadSolicitante.objects.all().order_by('id')
    unidades = aplicar_filtro_unidades(unidades, busqueda)

    paginator = Paginator(unidades, 1)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'unidades/listar.html', {
        'page_obj': page_obj,
        'busqueda': busqueda
    })

def detalle_unidad(request, id):
    unidad = get_object_or_404(UnidadSolicitante, id=id)

    return render(request, 'unidades/detalle.html', {
        'unidad': unidad
    })


def crear_unidad(request):
    form = UnidadSolicitanteForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, 'Unidad solicitante registrada correctamente.')
            return redirect('listar_unidades')
        notificar_errores_formulario(request, form)

    return render(request, 'unidades/formulario.html', {
        'form': form,
        'titulo': 'Registrar unidad solicitante'
    })


def editar_unidad(request, id):
    unidad = get_object_or_404(UnidadSolicitante, id=id)
    form = UnidadSolicitanteForm(request.POST or None, instance=unidad)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, 'Unidad solicitante actualizada correctamente.')
            return redirect('listar_unidades')
        notificar_errores_formulario(request, form)

    return render(request, 'unidades/formulario.html', {
        'form': form,
        'titulo': 'Editar unidad solicitante'
    })


def eliminar_unidad(request, id):
    unidad = get_object_or_404(UnidadSolicitante, id=id)

    if request.method == 'POST':
        # Validar si la unidad tiene reservas asociadas
        reservas_asociadas = ReservaMovil.objects.filter(unidad_solicitante=unidad)
        
        if reservas_asociadas.exists():
            messages.error(
                request, 
                f'No se puede eliminar la unidad solicitante "{unidad.nombre}" '
                f'porque tiene {reservas_asociadas.count()} reserva(s) de móvil asociada(s). '
                'Debe cancelar o finalizar todas las reservas antes de eliminar.'
            )
            return redirect('listar_unidades')
        
        unidad.delete()
        messages.success(request, 'Unidad solicitante eliminada correctamente.')
        return redirect('listar_unidades')

    return redirect('listar_unidades')


# -------------------------
# CRUD ACTIVIDAD DE SALUD
# -------------------------

def listar_actividades(request):
    busqueda = request.GET.get('buscar', '')

    actividades = ActividadSalud.objects.all().order_by('id')
    actividades = aplicar_filtro_actividades(actividades, busqueda)

    paginator = Paginator(actividades, 1)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'actividades/listar.html', {
        'page_obj': page_obj,
        'busqueda': busqueda
    })

def detalle_actividad(request, id):
    actividad = get_object_or_404(ActividadSalud, id=id)

    return render(request, 'actividades/detalle.html', {
        'actividad': actividad
    })


def crear_actividad(request):
    form = ActividadSaludForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, 'Actividad registrada correctamente.')
            return redirect('listar_actividades')
        notificar_errores_formulario(request, form)

    return render(request, 'actividades/formulario.html', {
        'form': form,
        'titulo': 'Registrar actividad de salud'
    })


def editar_actividad(request, id):
    actividad = get_object_or_404(ActividadSalud, id=id)
    form = ActividadSaludForm(request.POST or None, instance=actividad)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, 'Actividad actualizada correctamente.')
            return redirect('listar_actividades')
        notificar_errores_formulario(request, form)

    return render(request, 'actividades/formulario.html', {
        'form': form,
        'titulo': 'Editar actividad de salud'
    })


def eliminar_actividad(request, id):
    actividad = get_object_or_404(ActividadSalud, id=id)

    if request.method == 'POST':
        # Validar si la actividad tiene reservas asociadas
        reservas_asociadas = ReservaMovil.objects.filter(actividad=actividad)
        
        if reservas_asociadas.exists():
            messages.error(
                request, 
                f'No se puede eliminar la actividad "{actividad.nombre}" '
                f'porque tiene {reservas_asociadas.count()} reserva(s) de móvil asociada(s). '
                'Debe cancelar o finalizar todas las reservas antes de eliminar.'
            )
            return redirect('listar_actividades')
        
        actividad.delete()
        messages.success(request, 'Actividad eliminada correctamente.')
        return redirect('listar_actividades')

    return redirect('listar_actividades')


# -------------------------
# CRUD DESTINO
# -------------------------

def listar_destinos(request):
    busqueda = request.GET.get('buscar', '')

    destinos = Destino.objects.all().order_by('id')
    destinos = aplicar_filtro_destinos(destinos, busqueda)

    paginator = Paginator(destinos, 1)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'destinos/listar.html', {
        'page_obj': page_obj,
        'busqueda': busqueda
    })

def detalle_destino(request, id):
    destino = get_object_or_404(Destino, id=id)

    return render(request, 'destinos/detalle.html', {
        'destino': destino
    })


def crear_destino(request):
    form = DestinoForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, 'Destino registrado correctamente.')
            return redirect('listar_destinos')
        notificar_errores_formulario(request, form)

    return render(request, 'destinos/formulario.html', {
        'form': form,
        'titulo': 'Registrar destino'
    })


def editar_destino(request, id):
    destino = get_object_or_404(Destino, id=id)
    form = DestinoForm(request.POST or None, instance=destino)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, 'Destino actualizado correctamente.')
            return redirect('listar_destinos')
        notificar_errores_formulario(request, form)

    return render(request, 'destinos/formulario.html', {
        'form': form,
        'titulo': 'Editar destino'
    })


def eliminar_destino(request, id):
    destino = get_object_or_404(Destino, id=id)

    if request.method == 'POST':
        # Validar si el destino tiene reservas asociadas
        reservas_asociadas = ReservaMovil.objects.filter(destino=destino)
        
        if reservas_asociadas.exists():
            messages.error(
                request, 
                f'No se puede eliminar el destino "{destino.nombre_lugar}" '
                f'porque tiene {reservas_asociadas.count()} reserva(s) de móvil asociada(s). '
                'Debe cancelar o finalizar todas las reservas antes de eliminar.'
            )
            return redirect('listar_destinos')
        
        destino.delete()
        messages.success(request, 'Destino eliminado correctamente.')
        return redirect('listar_destinos')

    return redirect('listar_destinos')


# -------------------------
# CRUD RESERVA MOVIL
# -------------------------

def listar_reservas(request):
    busqueda = request.GET.get('buscar', '')

    reservas = ReservaMovil.objects.all().order_by('-fecha_reserva')
    reservas = aplicar_filtro_reservas(reservas, busqueda)

    paginator = Paginator(reservas, 1)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'reservas/listar.html', {
        'page_obj': page_obj,
        'busqueda': busqueda
    })

def detalle_reserva(request, id):
    reserva = get_object_or_404(ReservaMovil, id=id)

    return render(request, 'reservas/detalle.html', {
        'reserva': reserva
    })


def crear_reserva(request):
    if not Vehiculo.objects.exists():
        messages.warning(request, 'Debe registrar al menos un vehículo antes de crear una reserva.')
        return redirect('listar_vehiculos')

    if not Conductor.objects.exists():
        messages.warning(request, 'Debe registrar al menos un conductor antes de crear una reserva.')
        return redirect('listar_conductores')

    if not UnidadSolicitante.objects.exists():
        messages.warning(request, 'Debe registrar al menos una unidad solicitante antes de crear una reserva.')
        return redirect('listar_unidades')

    if not ActividadSalud.objects.exists():
        messages.warning(request, 'Debe registrar al menos una actividad de salud antes de crear una reserva.')
        return redirect('listar_actividades')

    if not Destino.objects.exists():
        messages.warning(request, 'Debe registrar al menos un destino antes de crear una reserva.')
        return redirect('listar_destinos')

    form = ReservaMovilForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, 'Reserva registrada correctamente.')
            return redirect('listar_reservas')
        notificar_errores_formulario(request, form)

    return render(request, 'reservas/formulario.html', {
        'form': form,
        'titulo': 'Registrar reserva de móvil'
    })


def editar_reserva(request, id):
    reserva = get_object_or_404(ReservaMovil, id=id)
    form = ReservaMovilForm(request.POST or None, instance=reserva)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, 'Reserva actualizada correctamente.')
            return redirect('listar_reservas')
        notificar_errores_formulario(request, form)

    return render(request, 'reservas/formulario.html', {
        'form': form,
        'titulo': 'Editar reserva de móvil'
    })


def eliminar_reserva(request, id):
    reserva = get_object_or_404(ReservaMovil, id=id)

    if request.method == 'POST':
        reserva.delete()
        messages.success(request, 'Reserva eliminada correctamente.')
        return redirect('listar_reservas')

    return redirect('listar_reservas')


def exportar_vehiculos_excel(request):
    busqueda = request.GET.get('buscar', '')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehículos"

    ws.append(['Patente', 'Tipo', 'Marca', 'Modelo', 'Año', 'Capacidad', 'Estado'])

    vehiculos = Vehiculo.objects.all()
    vehiculos = aplicar_filtro_vehiculos(vehiculos, busqueda)

    for vehiculo in vehiculos:
        ws.append([
            vehiculo.patente,
            vehiculo.tipo,
            vehiculo.marca,
            vehiculo.modelo,
            vehiculo.anio,
            vehiculo.capacidad,
            vehiculo.estado
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=vehiculos.xlsx'

    wb.save(response)
    return response


def exportar_vehiculos_pdf(request):
    busqueda = request.GET.get('buscar', '')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=vehiculos.pdf'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elementos = []
    estilos = getSampleStyleSheet()

    titulo = Paragraph("Reporte de Vehículos", estilos['Title'])
    fecha = Paragraph(f"Fecha de generación: {datetime.now().strftime('%d-%m-%Y %H:%M')}", estilos['Normal'])

    elementos.append(titulo)
    elementos.append(Spacer(1, 12))
    elementos.append(fecha)
    elementos.append(Spacer(1, 12))

    data = [['Patente', 'Tipo', 'Marca', 'Modelo', 'Año', 'Estado']]

    vehiculos = Vehiculo.objects.all()
    vehiculos = aplicar_filtro_vehiculos(vehiculos, busqueda)

    for vehiculo in vehiculos:
        data.append([
            vehiculo.patente,
            vehiculo.tipo,
            vehiculo.marca,
            vehiculo.modelo,
            str(vehiculo.anio),
            vehiculo.estado
        ])

    tabla = Table(data)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))

    elementos.append(tabla)
    doc.build(elementos)

    return response

def exportar_conductores_excel(request):
    busqueda = request.GET.get('buscar', '')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Conductores"

    ws.append(['Nombre completo', 'RUT', 'Teléfono', 'Correo', 'Tipo licencia', 'Estado'])

    conductores = Conductor.objects.all()
    conductores = aplicar_filtro_conductores(conductores, busqueda)

    for conductor in conductores:
        ws.append([
            conductor.nombre_completo,
            conductor.rut,
            conductor.telefono,
            conductor.correo,
            conductor.tipo_licencia,
            conductor.estado
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=conductores.xlsx'

    wb.save(response)
    return response


def exportar_conductores_pdf(request):
    busqueda = request.GET.get('buscar', '')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=conductores.pdf'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elementos = []
    estilos = getSampleStyleSheet()

    elementos.append(Paragraph("Reporte de Conductores", estilos['Title']))
    elementos.append(Spacer(1, 12))
    elementos.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%d-%m-%Y %H:%M')}", estilos['Normal']))
    elementos.append(Spacer(1, 12))

    data = [['Nombre', 'RUT', 'Teléfono', 'Licencia', 'Estado']]

    conductores = Conductor.objects.all()
    conductores = aplicar_filtro_conductores(conductores, busqueda)

    for conductor in conductores:
        data.append([
            conductor.nombre_completo,
            conductor.rut,
            conductor.telefono,
            conductor.tipo_licencia,
            conductor.estado
        ])

    tabla = Table(data)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))

    elementos.append(tabla)
    doc.build(elementos)

    return response

def exportar_unidades_excel(request):
    busqueda = request.GET.get('buscar', '')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Unidades"

    ws.append(['Nombre', 'Responsable', 'Teléfono contacto', 'Correo contacto', 'Descripción'])

    unidades = UnidadSolicitante.objects.all()
    unidades = aplicar_filtro_unidades(unidades, busqueda)

    for unidad in unidades:
        ws.append([
            unidad.nombre,
            unidad.responsable,
            unidad.telefono_contacto,
            unidad.correo_contacto,
            unidad.descripcion
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=unidades_solicitantes.xlsx'

    wb.save(response)
    return response


def exportar_unidades_pdf(request):
    busqueda = request.GET.get('buscar', '')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=unidades_solicitantes.pdf'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elementos = []
    estilos = getSampleStyleSheet()

    elementos.append(Paragraph("Reporte de Unidades Solicitantes", estilos['Title']))
    elementos.append(Spacer(1, 12))
    elementos.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%d-%m-%Y %H:%M')}", estilos['Normal']))
    elementos.append(Spacer(1, 12))

    data = [['Nombre', 'Responsable', 'Teléfono', 'Correo']]

    unidades = UnidadSolicitante.objects.all()
    unidades = aplicar_filtro_unidades(unidades, busqueda)

    for unidad in unidades:
        data.append([
            unidad.nombre,
            unidad.responsable,
            unidad.telefono_contacto,
            unidad.correo_contacto
        ])

    tabla = Table(data)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))

    elementos.append(tabla)
    doc.build(elementos)

    return response

def exportar_actividades_excel(request):
    busqueda = request.GET.get('buscar', '')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Actividades"

    ws.append(['Nombre', 'Descripción'])

    actividades = ActividadSalud.objects.all()
    actividades = aplicar_filtro_actividades(actividades, busqueda)

    for actividad in actividades:
        ws.append([
            actividad.nombre,
            actividad.descripcion
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=actividades_salud.xlsx'

    wb.save(response)
    return response


def exportar_actividades_pdf(request):
    busqueda = request.GET.get('buscar', '')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=actividades_salud.pdf'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elementos = []
    estilos = getSampleStyleSheet()

    elementos.append(Paragraph("Reporte de Actividades de Salud", estilos['Title']))
    elementos.append(Spacer(1, 12))
    elementos.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%d-%m-%Y %H:%M')}", estilos['Normal']))
    elementos.append(Spacer(1, 12))

    data = [['Nombre', 'Descripción']]

    actividades = ActividadSalud.objects.all()
    actividades = aplicar_filtro_actividades(actividades, busqueda)

    for actividad in actividades:
        data.append([
            actividad.nombre,
            actividad.descripcion
        ])

    tabla = Table(data)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))

    elementos.append(tabla)
    doc.build(elementos)

    return response

def exportar_destinos_excel(request):
    busqueda = request.GET.get('buscar', '')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Destinos"

    ws.append([
        'Lugar',
        'Dirección',
        'Referencia',
        'Latitud',
        'Longitud'
    ])

    destinos = Destino.objects.all()
    destinos = aplicar_filtro_destinos(destinos, busqueda)

    for destino in destinos:
        ws.append([
            destino.nombre_lugar,
            destino.direccion,
            destino.referencia,
            destino.latitud,
            destino.longitud
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename=destinos.xlsx'

    wb.save(response)

    return response


def exportar_destinos_pdf(request):
    busqueda = request.GET.get('buscar', '')
    
    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = 'attachment; filename=destinos.pdf'

    doc = SimpleDocTemplate(response, pagesize=letter)

    elementos = []

    estilos = getSampleStyleSheet()

    elementos.append(
        Paragraph("Reporte de Destinos", estilos['Title'])
    )

    elementos.append(Spacer(1, 12))

    elementos.append(
        Paragraph(
            f"Fecha de generación: {datetime.now().strftime('%d-%m-%Y %H:%M')}",
            estilos['Normal']
        )
    )

    elementos.append(Spacer(1, 12))

    data = [[
        'Lugar',
        'Dirección',
        'Referencia',
        'Latitud',
        'Longitud'
    ]]

    destinos = Destino.objects.all()
    destinos = aplicar_filtro_destinos(destinos, busqueda)

    for destino in destinos:
        data.append([
            destino.nombre_lugar,
            destino.direccion,
            destino.referencia,
            str(destino.latitud),
            str(destino.longitud)
        ])

    tabla = Table(data)

    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))

    elementos.append(tabla)

    doc.build(elementos)

    return response

def exportar_reservas_excel(request):
    busqueda = request.GET.get('buscar', '')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reservas"

    ws.append([
        'Fecha',
        'Hora inicio',
        'Hora término',
        'Vehículo',
        'Conductor',
        'Unidad solicitante',
        'Actividad',
        'Destino',
        'Estado',
        'Motivo'
    ])

    reservas = ReservaMovil.objects.all()
    reservas = aplicar_filtro_reservas(reservas, busqueda)

    for reserva in reservas:
        ws.append([
            reserva.fecha_reserva,
            reserva.hora_inicio,
            reserva.hora_termino,
            str(reserva.vehiculo),
            str(reserva.conductor),
            str(reserva.unidad_solicitante),
            str(reserva.actividad),
            str(reserva.destino),
            reserva.estado,
            reserva.motivo
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=reservas_moviles.xlsx'

    wb.save(response)
    return response


def exportar_reservas_pdf(request):
    busqueda = request.GET.get('buscar', '')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=reservas_moviles.pdf'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elementos = []
    estilos = getSampleStyleSheet()

    elementos.append(Paragraph("Reporte de Reservas de Móviles", estilos['Title']))
    elementos.append(Spacer(1, 12))
    elementos.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%d-%m-%Y %H:%M')}", estilos['Normal']))
    elementos.append(Spacer(1, 12))

    data = [[
        'Fecha',
        'Horario',
        'Vehículo',
        'Conductor',
        'Unidad',
        'Actividad',
        'Estado'
    ]]

    reservas = ReservaMovil.objects.all()
    reservas = aplicar_filtro_reservas(reservas, busqueda)

    for reserva in reservas:
        data.append([
            str(reserva.fecha_reserva),
            f"{reserva.hora_inicio} - {reserva.hora_termino}",
            str(reserva.vehiculo),
            str(reserva.conductor),
            str(reserva.unidad_solicitante),
            str(reserva.actividad),
            reserva.estado
        ])

    tabla = Table(data)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))

    elementos.append(tabla)
    doc.build(elementos)

    return response
